from dataclasses import dataclass
from io import BytesIO
import re

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.college import College
from app.models.course import Course

COURSE_IMPORT_HEADERS = (
    '学院', '课程名称', 'Slug', '别名', '分类', '学分', '描述', '川大课程号', '原开课单位',
)
SLUG_PATTERN = re.compile(r'^[a-z][a-z0-9]*(?:-[a-z0-9]+)*$')
MAX_IMPORT_ROWS = 5000
COURSE_CATEGORIES = {'通识', '专业必修', '专业选修', '实践'}


@dataclass(frozen=True)
class CourseImportRow:
    row_number: int
    college_name: str
    name: str
    slug: str
    aliases: list[str]
    category: str | None
    credit: float | None
    description: str | None


def parse_course_workbook(content: bytes) -> list[CourseImportRow]:
    if len(content) > 5 * 1024 * 1024:
        raise ValueError('Excel 文件不能超过 5MB')
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('无法读取 Excel 文件，请使用 .xlsx 格式') from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError('Excel 文件为空')
    normalized_header = tuple(str(value or '').strip() for value in header[:len(COURSE_IMPORT_HEADERS)])
    if normalized_header != COURSE_IMPORT_HEADERS:
        raise ValueError(f'表头必须为：{"、".join(COURSE_IMPORT_HEADERS)}')

    parsed: list[CourseImportRow] = []
    for row_number, values in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        if len(parsed) >= MAX_IMPORT_ROWS:
            raise ValueError(f'单次最多导入 {MAX_IMPORT_ROWS} 门课程')
        cells = list(values) + [None] * len(COURSE_IMPORT_HEADERS)
        aliases = [item.strip() for item in re.split(r'[;；]', str(cells[3] or '')) if item.strip()]
        credit = None
        if cells[5] is not None and str(cells[5]).strip():
            try:
                credit = float(cells[5])
            except (TypeError, ValueError) as exc:
                raise ValueError(f'第 {row_number} 行学分不是数字') from exc
        parsed.append(CourseImportRow(
            row_number=row_number,
            college_name=str(cells[0] or '').strip(),
            name=str(cells[1] or '').strip(),
            slug=str(cells[2] or '').strip(),
            aliases=aliases,
            category=str(cells[4]).strip() if cells[4] is not None and str(cells[4]).strip() else None,
            credit=credit,
            description=str(cells[6]).strip() if cells[6] is not None and str(cells[6]).strip() else None,
        ))
    if not parsed:
        raise ValueError('Excel 中没有课程数据')
    return parsed


async def import_courses(db: AsyncSession, rows: list[CourseImportRow], dry_run: bool) -> dict:
    colleges = list((await db.scalars(select(College))).all())
    courses = list((await db.scalars(select(Course))).all())
    college_by_name = {college.name: college for college in colleges}
    existing_names = {(str(course.college_id), course.name) for course in courses}
    existing_slugs = {course.slug for course in courses}
    batch_names: set[tuple[str, str]] = set()
    batch_slugs: set[str] = set()
    errors: list[dict] = []
    skipped: list[dict] = []
    pending: list[Course] = []

    for row in rows:
        college = college_by_name.get(row.college_name)
        row_errors = []
        if college is None:
            row_errors.append('学院不存在')
        if not row.name:
            row_errors.append('课程名称为空')
        if len(row.name) > 128:
            row_errors.append('课程名称超过 128 个字符')
        if not SLUG_PATTERN.fullmatch(row.slug):
            row_errors.append('Slug 必须是英文小写、数字和连字符，并以字母开头')
        if len(row.slug) > 128:
            row_errors.append('Slug 超过 128 个字符')
        if row.credit is not None and not 0 <= row.credit <= 30:
            row_errors.append('学分必须在 0 到 30 之间')
        if row.category is not None and row.category not in COURSE_CATEGORIES:
            row_errors.append('分类必须是通识、专业必修、专业选修或实践')
        if row.slug in existing_slugs or row.slug in batch_slugs:
            row_errors.append('Slug 已存在或在文件中重复')

        name_key = (str(college.id), row.name) if college else ('', row.name)
        if name_key in existing_names:
            skipped.append({'row': row.row_number, 'name': row.name, 'reason': '同学院课程已存在'})
            continue
        if name_key in batch_names:
            row_errors.append('同学院课程在文件中重复')
        if row_errors:
            errors.append({'row': row.row_number, 'name': row.name, 'messages': row_errors})
            continue

        batch_names.add(name_key)
        batch_slugs.add(row.slug)
        pending.append(Course(
            college_id=college.id,
            name=row.name,
            slug=row.slug,
            aliases=row.aliases,
            description=row.description,
            credit=row.credit,
            category=row.category,
        ))

    if not dry_run and errors:
        raise ValueError('Excel 校验未通过，请修正后重新导入')
    if not dry_run:
        db.add_all(pending)
        await db.flush()

    return {
        'total': len(rows),
        'ready': len(pending),
        'skipped': len(skipped),
        'error_count': len(errors),
        'errors': errors[:100],
        'skipped_items': skipped[:100],
        'imported': 0 if dry_run else len(pending),
    }
