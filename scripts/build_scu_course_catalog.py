import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
HEADERS = ['学院', '课程名称', 'Slug', '别名', '分类', '学分', '描述', '川大课程号', '原开课单位']

COLLEGE_NAMES = [
    '经济学院', '法学院', '文学与新闻学院', '外国语学院', '艺术学院', '历史文化学院（旅游学院）',
    '哲学系', '马克思主义学院', '体育学院', '公共管理学院', '商学院', '数学学院', '物理学院',
    '化学学院', '生命科学学院', '电子信息学院', '材料科学与工程学院', '机械工程学院',
    '电气工程学院', '计算机学院（软件学院、人工智能学院）', '建筑与环境学院', '水利水电学院',
    '化学工程学院', '轻工科学与工程学院', '高分子科学与工程学院', '空天科学与工程学院',
    '网络空间安全学院', '生物医学工程学院', '碳中和未来技术学院', '华西基础医学与法医学院',
    '华西临床医学院（华西医院）', '华西口腔医学院', '华西公共卫生学院', '华西药学院',
    '国际关系学院', '匹兹堡学院', '灾后重建与管理学院', '海外教育学院', '成人继续教育学院',
]

DIRECT_COLLEGE_MAP = {
    '历史文化学院（旅游学院、考古文博学院）': '历史文化学院（旅游学院）',
    '文学院': '文学与新闻学院',
    '新闻传播与出版学院': '文学与新闻学院',
    '古典学系': '历史文化学院（旅游学院）',
    '计算机学院': '计算机学院（软件学院、人工智能学院）',
    '软件学院': '计算机学院（软件学院、人工智能学院）',
    '人工智能学院': '计算机学院（软件学院、人工智能学院）',
    '计算机学院（软件学院、智能科学与技术学院）': '计算机学院（软件学院、人工智能学院）',
    '华西临床医学院': '华西临床医学院（华西医院）',
    'West China School of Medicine': '华西临床医学院（华西医院）',
    '碳中和技术创新班': '碳中和未来技术学院',
    '党委学生工作部（处）': '马克思主义学院',
    '党委学工部（军事教研室）': '马克思主义学院',
    '预科教育': '成人继续教育学院',
    '就业指导中心': '公共管理学院',
    '图书馆': '文学与新闻学院',
    '心理健康教育中心': '华西公共卫生学院',
    '生物治疗国家重点实验室': '生命科学学院',
    '分析测试中心': '化学学院',
    '实验室及设备管理处': '材料科学与工程学院',
    '工程训练中心': '机械工程学院',
    '工程设计中心': '机械工程学院',
    '电工电子中心': '电气工程学院',
    '电子实习中心': '电子信息学院',
    '化学基础实验教学中心': '化学学院',
    '计算机基础教学实验中心': '计算机学院（软件学院、人工智能学院）',
}

EXACT_SLUGS = {
    '中级计量经济学': 'intermediate-econometrics', '中级宏观经济学': 'intermediate-macro',
    '战略商业领袖（ACCA）': 'strategic-business-leader', '战略商业报告（ACCA）': 'strategic-business-report',
    '金融市场学': 'financial-markets', '金融计量学': 'financial-econometrics', '金融随机过程': 'financial-stochastics',
    '税法（ACCA）': 'tax-law-acca', '审计基础（ACCA先修课）': 'audit-foundations',
    '光学': 'basic-optics', '力学': 'mechanics', '电动力学': 'electrodynamics',
    '人文社科选修： 媒介素养（全英文）': 'pitt-media-literacy',
    '人文社科选修：媒介素养（全英文）': 'pitt-media-literacy',
    '晶体结构与衍射（全英文）': 'pitt-crystal-diffraction',
    '微加工与集成技术（全英文）': 'pitt-microfabrication',
    '人文社科选修8：电影流派（全英文）': 'pitt-film-genres',
    '高等数学': 'calculus', '线性代数': 'linear-algebra', '概率论与数理统计': 'probability-stats',
    '大学物理': 'college-physics', '大学化学': 'college-chemistry', '大学英语': 'college-english',
    '数据结构与算法': 'dsa', '数据结构': 'data-structures', '算法设计与分析': 'algorithm-design',
    '计算机组成原理': 'computer-architecture', '操作系统': 'operating-systems',
    '计算机网络': 'computer-networks', '数据库系统': 'database-systems',
    '编译原理': 'compiler-design', '软件工程': 'software-engineering',
    '离散数学': 'discrete-math', '数字逻辑': 'digital-logic', '人工智能': 'ai',
    '机器学习': 'machine-learning', '深度学习': 'deep-learning', '微观经济学': 'microeconomics',
    '宏观经济学': 'macroeconomics', '计量经济学': 'econometrics', '金融学': 'finance',
    '管理学': 'management', '会计学': 'accounting', '市场营销': 'marketing',
    '民法': 'civil-law', '刑法': 'criminal-law', '宪法': 'constitutional-law',
    '生物化学': 'biochemistry', '分子生物学': 'molecular-biology', '生理学': 'physiology',
    '病理学': 'pathology', '药理学': 'pharmacology', '诊断学': 'diagnostics',
}

SEMANTIC_TERMS = OrderedDict(sorted({
    '主题性创作': 'thematic-creation', '人体写生': 'figure-study', '油画人物': 'oil-figure',
    '文字学概要': 'linguistics-overview', '篆书创作': 'seal-script', '小组工作': 'group-work',
    '中国画论': 'chinese-painting-theory', '人体形态学': 'human-morphology', '人文经典导读': 'humanities-classics',
    '数学': 'mathematics', '法的智慧': 'legal-wisdom', '数据思维': 'data-thinking', '中华法系': 'chinese-legal-tradition',
    '精神应激': 'psychological-stress', '机电一体化': 'mechatronics', '经典影片': 'film-classics',
    '专业外语': 'technical-english', '生物大分子': 'biomacromolecules', '科学研究': 'scientific-research',
    '室内设计': 'interior-design', '公共空间': 'public-space', '品牌形象': 'brand-identity', '空间媒体': 'spatial-media',
    '交互媒体': 'interactive-media', '标志设计': 'logo-design', '装饰图案': 'decorative-patterns',
    '字体设计': 'type-design', '广告设计': 'advertising-design', '色彩基础': 'color-foundations',
    '素描基础': 'drawing-foundations', '古典绘画': 'classical-painting', '舞台编剧': 'stage-writing',
    '甲骨文字': 'oracle-bone-script', '美术简史': 'art-history', '弦歌': 'music-listening',
    '专项教学': 'specialized-coaching', '数字广告': 'digital-advertising', '社会研究': 'social-research',
    '生命奥秘': 'life-mysteries', '减灾服务': 'disaster-relief', '公益领袖': 'civic-leadership',
    '造物论': 'making-theory', '营建思想': 'building-thought', '图形创意': 'graphic-ideas',
    '花鸟画': 'bird-flower-painting', '人物画': 'figure-painting', '重彩山水': 'color-landscape',
    '写意花鸟': 'freehand-bird-flower', '写意人物': 'freehand-figure', '隶书': 'clerical-script',
    '楷书': 'regular-script', '草书': 'cursive-script', '行书': 'running-script', '书法': 'calligraphy',
    '素描': 'drawing', '声乐': 'vocal', '合唱': 'choral', '视唱练耳': 'ear-training', '伴奏': 'accompaniment',
    '钢琴': 'piano', '教学剧目': 'teaching-repertoire', '导演技巧': 'directing', '芭蕾': 'ballet',
    '古典基本功': 'classical-technique', '古典身韵': 'classical-movement', '赛事': 'sports-events',
    '天人之道': 'human-nature', '大学生心理': 'student-mental-health', '高等代数': 'advanced-algebra',
    '古代汉语': 'classical-chinese', '计算思维': 'computational-thinking', '国际公法': 'public-international-law',
    '中医学': 'traditional-medicine', '病案': 'case-analysis', '美容': 'medical-aesthetics',
    '抽象代数': 'abstract-algebra', '生物质加工': 'biomass-processing', '生物分离': 'bioseparation',
    '时装': 'fashion', '人文地理': 'human-geography', '思想品德': 'ethics-education', '心理健康': 'mental-health',
    '人因工程': 'human-factors', '全球气候': 'global-climate', '医患沟通': 'patient-communication',
    '眼病': 'eye-disease', '妇幼健康': 'maternal-child-health', '生命关怀': 'care-ethics',
    '唐宋诗文': 'tang-song-literature', '远离毒品': 'drug-prevention', '太极拳': 'tai-chi',
    '人力资源': 'human-resources', '生物质制造': 'biomanufacturing', '内在的宇宙': 'inner-universe',
    '信息素养': 'information-literacy', '问天寻梦': 'space-dreams', '坚韧': 'resilience',
    '劳动法': 'labor-law', '卫生政策': 'health-policy', '博弈论': 'game-theory', '外文专业': 'foreign-literature',
    '趣味医学': 'medical-insights', '结构选型': 'structural-selection', '博物学': 'natural-history',
    '智慧城市': 'smart-cities', '税务合规': 'tax-compliance', '俄译汉': 'russian-chinese-translation',
    '波兰': 'poland-studies', '俄英复语': 'russian-english', '法汉': 'french-chinese-translation',
    '夺桥徒搏': 'bridge-combat', '教育学': 'pedagogy', '教练员': 'coaching', '卓越工程师': 'engineering-excellence',
    '名师讲堂': 'master-lectures', '科研训练': 'research-training', '科研轮训': 'research-rotation',
    '山水画': 'landscape-painting', '全球气候变化': 'climate-change', '安全': 'safety', '危机': 'crisis',
    '数据资源': 'data-resources', '社会组织': 'social-organizations', '金融随机': 'financial-stochastics',
    '金融计量': 'financial-econometrics', '等离子体': 'plasma', '量子光学': 'quantum-optics',
    '传感物理': 'sensor-physics', '同步辐射': 'synchrotron', '品牌管理': 'brand-management',
    '战略商业': 'strategic-business', '先进材料': 'advanced-materials', '当代中国经济': 'contemporary-china-economy',
    '服务运作': 'service-operations', '粒子物理': 'particle-physics', '纳米材料': 'nanomaterials',
    '国民经济': 'national-economy', '加速器': 'accelerator', '激光原理': 'laser-principles',
    '物联网': 'internet-of-things', '计算机通信': 'computer-communications', '随机信号': 'random-signals',
    '时尚': 'fashion', '医用材料': 'biomedical-materials', '设计经典': 'design-classics',
    '物理学专业': 'physics-major', '超声断面': 'ultrasound-sections', '断面解剖': 'sectional-anatomy',
    '税法': 'tax-law', '审计': 'auditing', '战略管理': 'strategic-management', '财务会计': 'financial-accounting',
    '业绩管理': 'performance-management', '雷达': 'radar', '线性规划': 'linear-programming',
    '法的精神': 'legal-philosophy', '古希腊': 'greco-roman', '社会设计': 'social-design',
    '病理生理': 'pathophysiology', '药理': 'pharmacology', '推进': 'propulsion', '世界当代史': 'contemporary-history',
    '数字信号': 'digital-signal', '数学物理': 'mathematical-physics', '仲裁': 'arbitration', '立法': 'legislation',
    '法律实证': 'empirical-law', '科学数据': 'scientific-data', '化妆品': 'cosmetics', '新生研讨': 'freshman-seminar',
    '口译': 'interpreting', '笔译': 'translation', '日语': 'japanese', '法语': 'french', '俄语': 'russian',
    '法国社会': 'french-society', '英语外来语': 'loanwords', '报刊': 'press-reading', '词汇学': 'lexicology',
    '核与辐射': 'nuclear-radiation', '剂量防护': 'radiation-protection', '核分析': 'nuclear-analysis',
    '药物化学': 'medicinal-chemistry', '机能学': 'functional-medicine', '匹克球': 'pickleball',
    '国防体育': 'defense-sports', '运动员': 'athlete-development', '生产实习': 'industrial-internship',
    '妇产科': 'obstetrics', '全科医学': 'general-practice', '神经内科': 'neurology', '外科实习': 'surgical-internship',
    '精神科': 'psychiatry', '儿科实习': 'pediatric-internship', '内科实习': 'medical-internship',
    '预防医学': 'preventive-medicine', '跨国并购': 'cross-border-mergers', '专项教学': 'specialized-coaching',
    '时间序列': 'time-series', '财政': 'public-finance', '税收': 'taxation', '税费': 'taxation',
    '结算': 'settlement', '预算': 'budget', '跨境': 'cross-border', '电商': 'ecommerce',
    '筹划': 'planning', '决策': 'decision-making', '资本论': 'capital', '国际税': 'intl-tax',
    '人格权': 'personality-rights', '侵权': 'tort', '公司法': 'corporate-law', '合同法': 'contract-law',
    '公民': 'civic', '国家安全': 'national-security', '法理': 'jurisprudence', '法治': 'rule-of-law',
    '社会调查': 'social-survey', '说文解字': 'etymology', '古代文论': 'literary-theory',
    '古典文献': 'classics', '古文字': 'paleography', '符号学': 'semiotics', '国情': 'country-studies',
    '大学语文': 'chinese-language', '广告摄影': 'ad-photography', '广告文案': 'ad-copy',
    '广告策划': 'ad-planning', '新媒体': 'new-media', '短视频': 'short-video', '视听': 'audiovisual',
    '节目': 'programming', '影视': 'film-tv', '纪录片': 'documentary', '普通话': 'mandarin',
    '汉字': 'chinese-characters', '现代汉语': 'modern-chinese', '第二语言': 'second-language',
    '编辑': 'editing', '人类学': 'anthropology', '养生': 'wellness', '诗歌': 'poetry',
    '莎士比亚': 'shakespeare', '语言与社会': 'sociolinguistics', '韩语': 'korean', '古希腊语': 'ancient-greek',
    '拉丁语': 'latin', '儒家': 'confucianism', '四书': 'four-books', '道家': 'daoism', '庄': 'daoism',
    '公羊传': 'gongyang-commentary', '穀梁传': 'guliang-commentary', '诗经': 'classic-poetry',
    '上古汉语': 'archaic-chinese', '青铜器': 'bronze-culture', '丝绸之路': 'silk-road',
    '民族': 'ethnicity', '碑铭': 'epigraphy', '纸草': 'papyrology', '先秦': 'pre-qin',
    '形而上学': 'metaphysics', '精神分析': 'psychoanalysis', '批判性思维': 'critical-thinking',
    '美学': 'aesthetics', '女性学': 'gender-studies', '形势与政策': 'current-affairs',
    '毛泽东思想': 'mao-thought', '社会主义': 'socialism', '国家战略': 'strategy',
    '乒乓球': 'table-tennis', '排球': 'volleyball', '篮球': 'basketball', '网球': 'tennis',
    '游泳': 'swimming', '蛙泳': 'breaststroke', '武术': 'martial-arts', '射箭': 'archery',
    '瑜伽': 'yoga', '田径': 'athletics', '体育': 'physical-education', '定向': 'orienteering',
    '社会工作': 'social-work', '劳动关系': 'labor-relations', '社会福利': 'social-welfare',
    '房地产': 'real-estate', '用户研究': 'user-research', '信息组织': 'information-organization',
    '智慧办公': 'digital-office', '组织行为': 'organizational-behavior', '用户画像': 'user-profiling',
    '商业文本': 'business-text', '消费者行为': 'consumer-behavior', '市场调查': 'market-research',
    '领导力': 'leadership', '可靠性': 'reliability', '系统工程': 'systems-engineering',
    '数学规划': 'mathematical-programming', '图与网络': 'graph-networks', '回归分析': 'regression',
    '代数学': 'algebra', '交换代数': 'commutative-algebra', '同调代数': 'homological-algebra',
    '微分流形': 'differential-manifolds', '泛函分析': 'functional-analysis', '矩阵论': 'matrix-theory',
    '数值分析': 'numerical-analysis', '解析几何': 'analytic-geometry', '微分几何': 'differential-geometry',
    '椭圆型方程': 'elliptic-equations', '积分方程': 'integral-equations', '并行计算': 'parallel-computing',
    '概率统计': 'probability-stats', '统计软件': 'statistical-software', '力学': 'mechanics',
    '电动力学': 'electrodynamics', '广义相对论': 'general-relativity', '宇宙学': 'cosmology',
    '量子场论': 'quantum-field-theory', '群论': 'group-theory', '热工': 'thermal-engineering',
    '仪器分析': 'instrumental-analysis', '有机反应': 'organic-reactions', '谱图': 'spectroscopy',
    '核磁共振': 'nmr', '阻燃': 'flame-retardancy', '园林': 'landscape', '遗传': 'genetics',
    '优生': 'eugenics', '内分泌': 'endocrinology', '发酵': 'fermentation', '土壤': 'soil-science',
    '基因工程': 'genetic-engineering', '生物制品': 'biologics', '细胞工程': 'cell-engineering',
    '蛋白质结构': 'protein-structure', '酶工程': 'enzyme-engineering', '光谱': 'spectroscopy',
    '多模态': 'multimodal', '显示技术': 'display-tech', '嵌入式': 'embedded-systems',
    '传感器': 'sensors', '增材制造': 'additive-manufacturing', '微机电': 'mems', '热流体': 'thermofluids',
    '表面工程': 'surface-engineering', '电机': 'electrical-machines', '电力设备': 'power-equipment',
    '脑机接口': 'brain-computer-interface', '云和云AI': 'cloud-ai', '图论': 'graph-theory',
    '数值计算': 'scientific-computing', '数据挖掘': 'data-mining', '汇编': 'assembly',
    '图形学': 'computer-graphics', '体系结构': 'computer-architecture', '问题求解': 'problem-solving',
    '固体废物': 'solid-waste', '垃圾处理': 'waste-management', '大气污染': 'air-pollution',
    '断裂力学': 'fracture-mechanics', '弹性力学': 'elasticity', '排水工程': 'drainage-engineering',
    '桥梁': 'bridge-engineering', '水污染': 'water-pollution', '泵站': 'pump-stations',
    '水质': 'water-quality', '生物多样性': 'biodiversity', '水文': 'hydrology', '水轮机': 'hydroturbines',
    '灌溉': 'irrigation', '继电保护': 'relay-protection', '边坡': 'slope-stability', '钢筋混凝土': 'reinforced-concrete',
    '钢结构': 'steel-structures', '化工原理': 'chemical-engineering-principles', '制药': 'pharmaceutical',
    '传热': 'heat-transfer', '生物反应': 'bioreaction', '轻化': 'light-chemistry', '制革': 'leather-technology',
    '服装': 'apparel', '立体裁剪': 'draping', '高分子': 'polymer', '塑料': 'plastics',
    '航空发动机': 'aero-engines', '飞行力学': 'flight-dynamics', '飞行器': 'aircraft', '隐身': 'stealth',
    '恶意代码': 'malware', '网络攻防': 'cyber-defense', '保密检查': 'security-audit', '医疗器械': 'medical-devices',
    '生物力学': 'biomechanics', '肿瘤': 'oncology', '骨修复': 'bone-repair', '人体寄生虫': 'parasitology',
    '胚胎': 'embryology', '司法鉴定': 'forensic-identification', '法医': 'forensics', '解剖': 'anatomy',
    '临床': 'clinical', '内分泌': 'endocrinology', '消化': 'gastroenterology', '泌尿': 'urology',
    '神经科学': 'neuroscience', '皮肤病': 'dermatology', '听力': 'audiology', '营养': 'nutrition',
    '输血': 'transfusion', '康复': 'rehabilitation', '影像': 'medical-imaging', '急诊': 'emergency-medicine',
    '感染': 'infectious-disease', '呼吸': 'respiratory', '麻醉': 'anesthesiology', '眼科学': 'ophthalmology',
    '口腔': 'dentistry', '义齿': 'prosthodontics', '牙体': 'endodontics', '卫生检疫': 'health-quarantine',
    '毒理': 'toxicology', '流行病学': 'epidemiology', '药事': 'pharmacy-law', '新药': 'drug-development',
    '全球治理': 'global-governance', '国际冲突': 'international-conflict', '大国关系': 'great-power-relations',
    '中国特色社会主义': 'socialism-cn', '习近平新时代中国特色社会主义思想概论': 'xi-thought',
    '计算机程序设计': 'programming', '面向对象程序设计': 'oop', '程序设计': 'programming',
    '计算机系统': 'computer-systems', '计算机基础': 'computing-fund', '信息系统': 'info-systems',
    '数据科学': 'data-science', '大数据': 'big-data', '数据库': 'database', '数据分析': 'data-analytics',
    '数据结构': 'data-structures', '算法': 'algorithms', '网络安全': 'cybersecurity', '密码学': 'cryptography',
    '网络空间': 'cyberspace', '计算机网络': 'networks', '软件': 'software', '硬件': 'hardware',
    '人工智能': 'ai', '机器学习': 'ml', '深度学习': 'deep-learning', '智能': 'intelligence',
    '图像处理': 'image-processing', '计算机视觉': 'computer-vision', '自然语言处理': 'nlp',
    '嵌入式': 'embedded', '微电子': 'microelectronics', '通信': 'communications', '信号': 'signals',
    '电路': 'circuits', '电子': 'electronics', '电气': 'electrical', '自动控制': 'control-systems',
    '控制': 'control', '机器人': 'robotics', '机械': 'mechanical', '工程制图': 'eng-drawing',
    '工程力学': 'eng-mechanics', '材料力学': 'material-mechanics', '理论力学': 'theoretical-mechanics',
    '流体力学': 'fluid-mechanics', '热力学': 'thermodynamics', '量子力学': 'quantum-mechanics',
    '电磁场': 'electromagnetics', '光学': 'optics', '物理': 'physics', '化学': 'chemistry',
    '高等数学': 'calculus', '数学分析': 'math-analysis', '线性代数': 'linear-algebra',
    '概率论': 'probability', '数理统计': 'stats', '统计学': 'statistics', '运筹学': 'operations-research',
    '微积分': 'calculus', '微分方程': 'differential-equations', '离散数学': 'discrete-math',
    '金融': 'finance', '经济': 'economics', '会计': 'accounting', '财务': 'financial-mgmt',
    '投资': 'investment', '证券': 'securities', '保险': 'insurance', '贸易': 'trade', '营销': 'marketing',
    '管理': 'management', '供应链': 'supply-chain', '物流': 'logistics', '电子商务': 'ecommerce',
    '公共政策': 'public-policy', '公共管理': 'public-admin', '行政': 'administration',
    '国际关系': 'intl-relations', '政治': 'politics', '社会学': 'sociology', '心理学': 'psychology',
    '哲学': 'philosophy', '伦理学': 'ethics', '逻辑学': 'logic', '马克思主义': 'marxism',
    '法学': 'law', '法律': 'law', '民法': 'civil-law', '刑法': 'criminal-law', '宪法': 'constitutional-law',
    '诉讼法': 'procedure-law', '国际法': 'intl-law', '经济法': 'economic-law', '知识产权': 'ip-law',
    '中国文学': 'chinese-lit', '外国文学': 'world-lit', '文学': 'literature', '语言学': 'linguistics',
    '新闻': 'journalism', '传播': 'communications', '写作': 'writing', '英语': 'english', '日语': 'japanese',
    '法语': 'french', '德语': 'german', '俄语': 'russian', '西班牙语': 'spanish', '翻译': 'translation',
    '中国史': 'chinese-history', '世界史': 'world-history', '历史': 'history', '考古': 'archaeology',
    '旅游': 'tourism', '文化': 'culture', '艺术': 'arts', '美术': 'fine-arts', '音乐': 'music',
    '舞蹈': 'dance', '绘画': 'painting', '书法': 'calligraphy', '设计': 'design', '建筑': 'architecture',
    '城市规划': 'urban-planning', '环境': 'environment', '土木': 'civil-engineering', '水利': 'hydraulics',
    '水电': 'hydropower', '测量': 'surveying', '地质': 'geology', '能源': 'energy', '碳中和': 'carbon-neutrality',
    '材料': 'materials', '高分子': 'polymer', '化工': 'chemical-eng', '轻工': 'light-industry',
    '食品': 'food-science', '纺织': 'textiles', '生物医学工程': 'biomedical-eng', '生物工程': 'bioengineering',
    '生物技术': 'biotech', '生命科学': 'life-science', '生物学': 'biology', '生态学': 'ecology',
    '遗传学': 'genetics', '细胞生物学': 'cell-biology', '分子生物学': 'molecular-biology',
    '生物化学': 'biochemistry', '人体解剖学': 'anatomy', '解剖学': 'anatomy', '生理学': 'physiology',
    '病理学': 'pathology', '药理学': 'pharmacology', '免疫学': 'immunology', '微生物学': 'microbiology',
    '临床医学': 'clinical-medicine', '内科学': 'internal-medicine', '外科学': 'surgery', '儿科学': 'pediatrics',
    '妇产科学': 'obgyn', '神经病学': 'neurology', '精神病学': 'psychiatry', '医学影像': 'medical-imaging',
    '口腔': 'dentistry', '公共卫生': 'public-health', '流行病学': 'epidemiology', '卫生统计': 'health-stats',
    '药物': 'drug-science', '药学': 'pharmacy', '法医学': 'forensic-medicine', '护理': 'nursing',
    '体育': 'physical-education', '运动': 'sports', '军事': 'military', '创新创业': 'innovation',
    '创新': 'innovation', '创业': 'entrepreneurship', '职业生涯': 'career-planning', '就业': 'career',
    '实验': 'lab', '实习': 'internship', '实践': 'practice', '实训': 'training', '课程设计': 'course-design',
    '毕业设计': 'capstone', '毕业论文': 'thesis', '毕业实习': 'grad-internship', '论文写作': 'academic-writing',
    '研究方法': 'research-methods', '专题研究': 'special-topics', '专题': 'topics', '导论': 'intro',
    '概论': 'survey', '基础': 'fundamentals', '原理': 'principles', '理论': 'theory', '方法': 'methods',
    '高级': 'advanced', '中级': 'intermediate', '初级': 'elementary', '前沿': 'frontiers',
    '全英文': 'english-taught', '研讨课': 'seminar', '阅读': 'reading', '鉴赏': 'appreciation',
    '史': 'history', '应用': 'applications', '综合': 'integrated', '专题讲座': 'lectures',
}.items(), key=lambda item: len(item[0]), reverse=True))

DEPARTMENT_PREFIX = {
    '艺术学院': 'arts', '经济学院': 'econ', '法学院': 'law', '文学与新闻学院': 'lit', '外国语学院': 'lang',
    '历史文化学院（旅游学院）': 'history', '哲学系': 'phil', '马克思主义学院': 'marx', '体育学院': 'pe',
    '公共管理学院': 'pa', '商学院': 'biz', '数学学院': 'math', '物理学院': 'physics', '化学学院': 'chem',
    '生命科学学院': 'bio', '电子信息学院': 'ee', '材料科学与工程学院': 'mse', '机械工程学院': 'me',
    '电气工程学院': 'ee', '计算机学院（软件学院、人工智能学院）': 'cs', '建筑与环境学院': 'arch',
    '水利水电学院': 'hydro', '化学工程学院': 'che', '轻工科学与工程学院': 'light-eng',
    '高分子科学与工程学院': 'polymer', '空天科学与工程学院': 'aero', '网络空间安全学院': 'cyber',
    '生物医学工程学院': 'bme', '碳中和未来技术学院': 'carbon', '华西基础医学与法医学院': 'med',
    '华西临床医学院（华西医院）': 'clinical', '华西口腔医学院': 'dental', '华西公共卫生学院': 'ph',
    '华西药学院': 'pharm', '国际关系学院': 'ir', '匹兹堡学院': 'pitt', '灾后重建与管理学院': 'drm',
    '海外教育学院': 'intl-ed', '成人继续教育学院': 'continuing-ed',
}


def map_college(department: str, course_name: str) -> str:
    if department in COLLEGE_NAMES:
        return department
    if department in DIRECT_COLLEGE_MAP:
        return DIRECT_COLLEGE_MAP[department]
    keyword_targets = [
        (('计算机', '程序', '数据', '人工智能'), '计算机学院（软件学院、人工智能学院）'),
        (('化学', '分析测试'), '化学学院'), (('电路', '电子'), '电子信息学院'),
        (('工程', '机械', '制图'), '机械工程学院'), (('医学', '健康'), '华西公共卫生学院'),
        (('创新', '创业'), '商学院'), (('军事', '思想政治'), '马克思主义学院'),
    ]
    for keywords, target in keyword_targets:
        if any(keyword in course_name for keyword in keywords):
            return target
    return '公共管理学院'


def slugify_english(text: str) -> str:
    text = re.sub(r'([a-z])([A-Z])', r'\1-\2', text)
    return re.sub(r'-+', '-', re.sub(r'[^a-z0-9]+', '-', text.lower())).strip('-')


def semantic_slug(name: str, college: str, course_code: str, used: set[str]) -> str:
    cleaned = re.sub(r'[《》“”"\s]', '', name).strip()
    normalized = cleaned.replace('Ⅰ', 'i').replace('Ⅱ', 'ii').replace('Ⅲ', 'iii').replace('Ⅳ', 'iv')
    base = EXACT_SLUGS.get(cleaned) or EXACT_SLUGS.get(re.sub(r'[（(].*?[）)]', '', cleaned).strip())
    if not base and re.search(r'[A-Za-z]{2,}', cleaned):
        english = slugify_english(cleaned)
        if len(english) >= 3:
            base = english
    if not base:
        remaining = cleaned
        tokens = []
        for chinese, english in SEMANTIC_TERMS.items():
            if chinese in remaining:
                tokens.append(english)
                remaining = remaining.replace(chinese, ' ')
            if len(tokens) == 6:
                break
        if tokens:
            base = '-'.join(dict.fromkeys(tokens))
        else:
            raise ValueError(f'无法为课程设计语义 slug：{name}')
    base = slugify_english(base)[:110].rstrip('-')
    slug = base
    if slug in used:
        slug = f'{DEPARTMENT_PREFIX[college]}-{base}'
    if slug in used:
        qualifiers = []
        if re.search(r'全英文|英文', name): qualifiers.append('english-taught')
        if re.search(r'双语|中/英', name): qualifiers.append('bilingual')
        if re.search(r'课程设计|设计', name): qualifiers.append('design')
        if re.search(r'实验|试验', name): qualifiers.append('lab')
        if re.search(r'下|Ⅱ|2', name): qualifiers.append('part-two')
        if re.search(r'上|Ⅰ|1', name): qualifiers.append('part-one')
        level_match = re.search(r'(?:[-－]|第)?([1-9])(?:级|期|册|$)', name)
        if level_match:
            qualifiers.append(f'level-{level_match.group(1)}')
        if re.search(r'强基', name): qualifiers.append('honors')
        if re.search(r'试验班', name): qualifiers.append('pilot')
        if re.search(r'远古', name): qualifiers.append('ancient')
        if re.search(r'近现代', name): qualifiers.append('modern')
        if re.search(r'文艺复兴', name): qualifiers.append('renaissance')
        if re.search(r'ACCA', name, re.IGNORECASE): qualifiers.append('acca')
        for qualifier in qualifiers:
            candidate = f'{base}-{qualifier}'
            if candidate not in used:
                slug = candidate
                break
    if slug in used:
        for qualifier in ('fundamentals', 'theory', 'applications', 'practice'):
            candidate = f'{DEPARTMENT_PREFIX[college]}-{base}-{qualifier}'
            if candidate not in used:
                slug = candidate
                break
    if slug in used:
        raise ValueError(f'课程 slug 冲突且无法语义区分：{name}')
    used.add(slug)
    return slug


def category_for(record: dict) -> str:
    official = (record.get('kclbmc') or '').strip()
    name = record.get('kcm') or ''
    restriction = record.get('xkxzsm') or ''
    if official.startswith('通识') or official in {'文化素质公选课', '校级公共课'}:
        return '通识'
    if any(word in name for word in ('实验', '实习', '实训', '课程设计', '毕业设计')):
        return '实践'
    if '主修' in restriction:
        return '专业必修'
    return '专业选修'


def build(input_path: Path, output_path: Path) -> None:
    payload = json.loads(input_path.read_text(encoding='utf-8-sig'))
    records = payload['list']['records']
    unique = {}
    for record in records:
        unique.setdefault((record.get('kch'), record.get('kcm')), record)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '课程导入'
    sheet.append(HEADERS)
    used_slugs: set[str] = set()
    mapped_names: set[tuple[str, str]] = set()
    mapped_rows = []
    for record in unique.values():
        name = (record.get('kcm') or '').strip()
        code = (record.get('kch') or '').strip()
        department = (record.get('kkxsjc') or '').strip()
        if not name or not code:
            continue
        if '主修' not in (record.get('xkxzsm') or '') and name not in EXACT_SLUGS:
            continue
        college = map_college(department, name)
        if (college, name) in mapped_names:
            continue
        slug = semantic_slug(name, college, code, used_slugs)
        mapped_names.add((college, name))
        credit = float(record['xf']) if record.get('xf') not in (None, '') else None
        description = f'来源：四川大学教务处 2026-2027学年秋季课程目录；课程号：{code}。'
        mapped_rows.append([college, name, slug, '', category_for(record), credit, description, code, department])
    mapped_rows.sort(key=lambda row: (COLLEGE_NAMES.index(row[0]), row[1], row[7]))
    for row in mapped_rows:
        sheet.append(row)

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:I{sheet.max_row}'
    widths = [30, 38, 34, 24, 14, 10, 58, 18, 36]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    validation = DataValidation(type='list', formula1='"通识,专业必修,专业选修,实践"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f'E2:E{sheet.max_row}')

    guide = workbook.create_sheet('导入说明')
    guide.append(['字段', '说明'])
    guide_rows = [
        ('学院', '必须与后台现有学院名称完全一致，不会自动新增学院。'),
        ('课程名称', '同一学院内名称相同的课程会跳过。'),
        ('Slug', '必填；英文语义缩写，全局唯一，仅允许小写字母、数字和连字符。'),
        ('别名', '多个别名用英文分号分隔。'),
        ('分类', '通识、专业必修、专业选修或实践。'),
        ('学分', '0 到 30，可使用 0.5 学分。'),
        ('来源', '四川大学教务处公开课程安排页面。'),
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions['A'].width = 18
    guide.column_dimensions['B'].width = 90
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f'rows={len(mapped_rows)} output={output_path}')


if __name__ == '__main__':
    if len(sys.argv) != 3:
        raise SystemExit('usage: build_scu_course_catalog.py <source-json> <output-xlsx>')
    build(Path(sys.argv[1]), Path(sys.argv[2]))
